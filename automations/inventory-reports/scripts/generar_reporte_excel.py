"""
generar_reporte_excel.py
========================
Lee el Excel de stock de bodega y genera un nuevo Excel de reporte
con columnas de estado y colores, sin modificar el archivo original.

Estados asignados a cada fila:
  ✅ Actualizado        → SKU existe en Shopify y aparece UNA sola vez en el Excel
  ⏳ Pendiente          → SKU existe en Shopify pero está DUPLICADO en el Excel
                          (requiere decisión: ¿sumar? ¿usar el mayor?)
  ⚠️ Supervisión Humana → SKU NO existe en Shopify (no publicado, formato raro, etc.)

Uso:
  python generar_reporte_excel.py --excel "/ruta/al/Stock.xlsx"
  python generar_reporte_excel.py --excel "/ruta/al/Stock.xlsx" --output "Reporte_Stock.xlsx"
"""

import os
import time
import argparse
import requests
import pandas as pd
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

REPO_ROOT = Path(__file__).resolve().parents[3]
load_dotenv(REPO_ROOT / ".env")

SHOPIFY_SHOP_URL     = os.getenv("SHOPIFY_SHOP_URL", "")
SHOPIFY_ACCESS_TOKEN = os.getenv("SHOPIFY_ACCESS_TOKEN", "")
SHOPIFY_API_VERSION  = os.getenv("SHOPIFY_API_VERSION", "2024-04")

EXCEL_COL_SKU        = "Código"
EXCEL_COL_STOCK      = "Disponible"
REQUESTS_DELAY       = 0.4

# ── Colores ──────────────────────────────────────────────────────────────────
COLOR_GREEN_BG    = "C6EFCE"   # fondo verde suave
COLOR_GREEN_FT    = "276221"   # texto verde oscuro
COLOR_YELLOW_BG   = "FFEB9C"   # fondo amarillo
COLOR_YELLOW_FT   = "9C6500"   # texto amarillo oscuro
COLOR_RED_BG      = "FFC7CE"   # fondo rojo suave
COLOR_RED_FT      = "9C0006"   # texto rojo oscuro
COLOR_HEADER_BG   = "1F3864"   # azul oscuro header
COLOR_HEADER_FT   = "FFFFFF"   # blanco
COLOR_ALT_ROW     = "F2F2F2"   # gris muy claro para filas alternas
COLOR_SUMMARY_BG  = "2E75B6"   # azul medio para resumen
COLOR_SUMMARY_FT  = "FFFFFF"


# ── Helpers Shopify ───────────────────────────────────────────────────────────

def _shop_domain():
    return SHOPIFY_SHOP_URL.replace("https://", "").replace("http://", "").strip("/")

def _headers():
    return {
        "X-Shopify-Access-Token": SHOPIFY_ACCESS_TOKEN,
        "Content-Type": "application/json",
    }

def build_sku_map():
    """Devuelve {sku: inventory_item_id} de todas las variantes en Shopify."""
    print("🔍 Conectando con Shopify para obtener catálogo...")
    sku_map = {}
    url = f"https://{_shop_domain()}/admin/api/{SHOPIFY_API_VERSION}/products.json"
    params = {"limit": 250, "fields": "id,title,variants"}

    while url:
        resp = requests.get(url, headers=_headers(), params=params)
        resp.raise_for_status()
        for product in resp.json().get("products", []):
            for variant in product.get("variants", []):
                sku = (variant.get("sku") or "").strip()
                if sku:
                    sku_map[sku] = variant["inventory_item_id"]
        next_link = resp.links.get("next")
        url = next_link["url"] if next_link else None
        params = {}
        time.sleep(REQUESTS_DELAY)

    print(f"   → {len(sku_map)} variantes con SKU encontradas en Shopify.")
    return sku_map


# ── Clasificación ─────────────────────────────────────────────────────────────

def clasificar(df, sku_map):
    """Agrega las columnas Estado, Observación y Stock Sincronizado."""

    # Contar cuántas veces aparece cada SKU en el Excel
    conteo_sku = df[EXCEL_COL_SKU].value_counts()

    estados      = []
    observaciones = []
    stock_sync   = []

    for _, row in df.iterrows():
        sku   = row[EXCEL_COL_SKU]
        stock = int(row[EXCEL_COL_STOCK])
        veces = conteo_sku.get(sku, 1)

        if sku not in sku_map:
            # No existe en Shopify
            if " " in sku and not any(c.isdigit() for c in sku[:3]):
                obs = "SKU con espacios — verificar formato en Shopify"
            else:
                obs = "Producto no publicado en Shopify"
            estados.append("⚠️ Supervisión Humana")
            observaciones.append(obs)
            stock_sync.append("")

        elif veces > 1:
            # Existe en Shopify pero está duplicado en el Excel
            stock_total = int(df.loc[df[EXCEL_COL_SKU] == sku, EXCEL_COL_STOCK].sum())
            estados.append("⏳ Pendiente")
            observaciones.append(
                f"SKU duplicado en Excel ({veces} filas) — Stock total sumado: {stock_total}. "
                f"Confirmar cuál usar."
            )
            stock_sync.append(stock_total)

        else:
            # Único y en Shopify → listo
            estados.append("✅ Actualizado")
            observaciones.append("Sincronizado correctamente con Shopify")
            stock_sync.append(stock)

    df = df.copy()
    df["Estado"]             = estados
    df["Observación"]        = observaciones
    df["Stock Sincronizado"] = stock_sync
    return df


# ── Estilos helpers ───────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def _font(hex_color, bold=False, size=11):
    return Font(color=hex_color, bold=bold, size=size, name="Calibri")

def _border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


# ── Construcción del Excel ────────────────────────────────────────────────────

def build_excel(df, output_path, excel_path):
    wb = Workbook()

    # ── Hoja 1: Resumen ──────────────────────────────────────────────────────
    ws_resumen = wb.active
    ws_resumen.title = "📊 Resumen"

    total        = len(df)
    actualizados = (df["Estado"] == "✅ Actualizado").sum()
    pendientes   = (df["Estado"] == "⏳ Pendiente").sum()
    supervision  = (df["Estado"] == "⚠️ Supervisión Humana").sum()
    fecha        = datetime.now().strftime("%d/%m/%Y %H:%M")

    # Título
    ws_resumen.merge_cells("B2:F2")
    cell = ws_resumen["B2"]
    cell.value = "📦 Reporte de Sincronización de Stock — Zipp"
    cell.font  = Font(name="Calibri", bold=True, size=18, color=COLOR_HEADER_FT)
    cell.fill  = _fill(COLOR_HEADER_BG)
    cell.alignment = _center()
    ws_resumen.row_dimensions[2].height = 40

    ws_resumen.merge_cells("B3:F3")
    cell = ws_resumen["B3"]
    cell.value = f"Excel origen: {os.path.basename(excel_path)}   |   Generado: {fecha}"
    cell.font  = Font(name="Calibri", size=10, color="666666", italic=True)
    cell.alignment = _center()

    # Tarjetas de resumen
    tarjetas = [
        ("B5:C6", "Total Productos\nen Excel",    total,        COLOR_HEADER_BG, COLOR_HEADER_FT),
        ("D5:D6", "✅ Actualizados",              actualizados, COLOR_GREEN_BG,  COLOR_GREEN_FT),
        ("E5:E6", "⏳ Pendientes",                pendientes,   COLOR_YELLOW_BG, COLOR_YELLOW_FT),
        ("F5:F6", "⚠️ Supervisión\nHumana",       supervision,  COLOR_RED_BG,    COLOR_RED_FT),
    ]

    for cell_range, label, valor, bg, ft in tarjetas:
        ws_resumen.merge_cells(cell_range)
        start_cell = ws_resumen[cell_range.split(":")[0]]
        start_cell.value     = f"{label}\n{valor}"
        start_cell.font      = Font(name="Calibri", bold=True, size=14, color=ft)
        start_cell.fill      = _fill(bg)
        start_cell.alignment = _center()
        start_cell.border    = _border()
    ws_resumen.row_dimensions[5].height = 50
    ws_resumen.row_dimensions[6].height = 50

    # Tabla de categorías de no encontrados
    ws_resumen["B8"].value  = "📋 Detalle de Supervisión Humana"
    ws_resumen["B8"].font   = Font(name="Calibri", bold=True, size=12, color=COLOR_HEADER_BG)
    ws_resumen.merge_cells("B8:F8")

    headers_det = ["Causa", "Cantidad"]
    for i, h in enumerate(headers_det):
        c = ws_resumen.cell(row=9, column=2+i, value=h)
        c.font = _font(COLOR_HEADER_FT, bold=True)
        c.fill = _fill(COLOR_SUMMARY_BG)
        c.alignment = _center()
        c.border = _border()

    no_publicados = df[
        (df["Estado"] == "⚠️ Supervisión Humana") &
        (~df["Observación"].str.contains("espacios", na=False))
    ]
    sku_espacios = df[
        (df["Estado"] == "⚠️ Supervisión Humana") &
        (df["Observación"].str.contains("espacios", na=False))
    ]

    causas = [
        ("Producto no publicado en Shopify", len(no_publicados)),
        ("SKU con formato incorrecto (espacios)", len(sku_espacios)),
        ("SKU duplicado en Excel (sumar stocks)", pendientes),
    ]
    for r, (causa, cant) in enumerate(causas, start=10):
        c1 = ws_resumen.cell(row=r, column=2, value=causa)
        c2 = ws_resumen.cell(row=r, column=3, value=cant)
        for c in [c1, c2]:
            c.font = _font("333333")
            c.alignment = _center()
            c.border = _border()
            c.fill = _fill("F8F8F8" if r % 2 == 0 else "FFFFFF")

    # Instrucciones
    ws_resumen["B13"].value = "📌 Próximos pasos"
    ws_resumen["B13"].font  = Font(name="Calibri", bold=True, size=12, color=COLOR_HEADER_BG)
    ws_resumen.merge_cells("B13:F13")

    pasos = [
        "1. ✅ Los productos ACTUALIZADOS ya tienen su stock sincronizado en Shopify.",
        "2. ⏳ Los PENDIENTES tienen SKU duplicado en el Excel. Confirmar el stock correcto y aplicar manualmente.",
        "3. ⚠️ Los de SUPERVISIÓN HUMANA no existen en Shopify. Revisar si hay que publicarlos o corregir el SKU.",
    ]
    for i, paso in enumerate(pasos, start=14):
        c = ws_resumen.cell(row=i, column=2, value=paso)
        c.font = Font(name="Calibri", size=10, color="333333")
        ws_resumen.merge_cells(f"B{i}:F{i}")

    # Ancho de columnas del resumen
    ws_resumen.column_dimensions["A"].width = 3
    for col in ["B", "C", "D", "E", "F"]:
        ws_resumen.column_dimensions[col].width = 28

    # ── Hoja 2: Detalle Completo ─────────────────────────────────────────────
    ws_det = wb.create_sheet("📋 Detalle Completo")

    # Columnas a mostrar (originales + nuevas)
    cols_originales = [EXCEL_COL_SKU, "Descripción", "Stock Físico", "Asignado", EXCEL_COL_STOCK]
    cols_nuevas     = ["Estado", "Stock Sincronizado", "Observación"]
    cols_mostrar    = cols_nuevas + cols_originales

    # Header
    for col_idx, col_name in enumerate(cols_mostrar, start=1):
        c = ws_det.cell(row=1, column=col_idx, value=col_name)
        c.font      = _font(COLOR_HEADER_FT, bold=True, size=11)
        c.fill      = _fill(COLOR_HEADER_BG)
        c.alignment = _center()
        c.border    = _border()
    ws_det.row_dimensions[1].height = 30

    # Filas de datos
    for row_idx, (_, row) in enumerate(df[cols_mostrar].iterrows(), start=2):
        estado = row["Estado"]
        if "✅" in str(estado):
            bg, ft = COLOR_GREEN_BG,  COLOR_GREEN_FT
        elif "⏳" in str(estado):
            bg, ft = COLOR_YELLOW_BG, COLOR_YELLOW_FT
        else:
            bg, ft = COLOR_RED_BG,    COLOR_RED_FT

        for col_idx, col_name in enumerate(cols_mostrar, start=1):
            val = row[col_name]
            c = ws_det.cell(row=row_idx, column=col_idx, value=val)
            c.border = _border()
            c.font   = Font(name="Calibri", size=10)

            if col_name == "Estado":
                c.fill      = _fill(bg)
                c.font      = Font(name="Calibri", bold=True, size=10, color=ft)
                c.alignment = _center()
            elif col_idx % 2 == 0:
                c.fill      = _fill("F9F9F9")
                c.alignment = _left()
            else:
                c.fill      = _fill("FFFFFF")
                c.alignment = _left()

        ws_det.row_dimensions[row_idx].height = 20

    # Anchos de columna hoja detalle
    anchos = {
        "Estado": 22,
        "Stock Sincronizado": 18,
        "Observación": 55,
        EXCEL_COL_SKU: 30,
        "Descripción": 45,
        "Stock Físico": 13,
        "Asignado": 12,
        EXCEL_COL_STOCK: 13,
    }
    for col_idx, col_name in enumerate(cols_mostrar, start=1):
        ws_det.column_dimensions[get_column_letter(col_idx)].width = anchos.get(col_name, 18)

    # Freeze panes
    ws_det.freeze_panes = "A2"

    # ── Hoja 3: Solo Pendientes / Supervisión ────────────────────────────────
    ws_pend = wb.create_sheet("⚠️ Requieren Acción")

    df_accion = df[df["Estado"] != "✅ Actualizado"].reset_index(drop=True)

    for col_idx, col_name in enumerate(cols_mostrar, start=1):
        c = ws_pend.cell(row=1, column=col_idx, value=col_name)
        c.font      = _font(COLOR_HEADER_FT, bold=True)
        c.fill      = _fill(COLOR_HEADER_BG)
        c.alignment = _center()
        c.border    = _border()
    ws_pend.row_dimensions[1].height = 30

    for row_idx, (_, row) in enumerate(df_accion[cols_mostrar].iterrows(), start=2):
        estado = row["Estado"]
        if "⏳" in str(estado):
            bg, ft = COLOR_YELLOW_BG, COLOR_YELLOW_FT
        else:
            bg, ft = COLOR_RED_BG, COLOR_RED_FT

        for col_idx, col_name in enumerate(cols_mostrar, start=1):
            val = row[col_name]
            c = ws_pend.cell(row=row_idx, column=col_idx, value=val)
            c.border = _border()
            c.font   = Font(name="Calibri", size=10)

            if col_name == "Estado":
                c.fill      = _fill(bg)
                c.font      = Font(name="Calibri", bold=True, size=10, color=ft)
                c.alignment = _center()
            else:
                c.fill      = _fill("FFFFFF")
                c.alignment = _left()

        ws_pend.row_dimensions[row_idx].height = 20

    for col_idx, col_name in enumerate(cols_mostrar, start=1):
        ws_pend.column_dimensions[get_column_letter(col_idx)].width = anchos.get(col_name, 18)

    ws_pend.freeze_panes = "A2"

    wb.save(output_path)
    print(f"\n✅ Reporte guardado en: {output_path}")


# ── Main ──────────────────────────────────────────────────────────────────────

def run(excel_path, output_path):
    if not SHOPIFY_SHOP_URL or not SHOPIFY_ACCESS_TOKEN:
        raise ValueError("Faltan SHOPIFY_SHOP_URL o SHOPIFY_ACCESS_TOKEN en el .env")

    print(f"\n{'='*60}")
    print(f"  Generador de Reporte Excel — Stock Zipp")
    print(f"{'='*60}\n")

    # 1. Obtener catálogo Shopify
    sku_map = build_sku_map()

    # 2. Leer Excel
    print(f"\n📂 Leyendo Excel: {excel_path}")
    df = pd.read_excel(excel_path, sheet_name=0)
    df[EXCEL_COL_SKU]   = df[EXCEL_COL_SKU].astype(str).str.strip()
    df[EXCEL_COL_STOCK] = pd.to_numeric(df[EXCEL_COL_STOCK], errors="coerce").fillna(0).astype(int)
    df = df[df[EXCEL_COL_SKU].notna() & (df[EXCEL_COL_SKU] != "") & (df[EXCEL_COL_SKU] != "nan")]
    print(f"   → {len(df)} filas cargadas.")

    # 3. Clasificar
    print("\n🏷️  Clasificando productos...")
    df = clasificar(df, sku_map)

    actualizados = (df["Estado"] == "✅ Actualizado").sum()
    pendientes   = (df["Estado"] == "⏳ Pendiente").sum()
    supervision  = (df["Estado"] == "⚠️ Supervisión Humana").sum()

    print(f"   ✅ Actualizados       : {actualizados}")
    print(f"   ⏳ Pendientes         : {pendientes}")
    print(f"   ⚠️  Supervisión Humana : {supervision}")

    # 4. Construir Excel
    print(f"\n📊 Generando Excel de reporte...")
    build_excel(df, output_path, excel_path)

    print(f"\n{'='*60}")
    print(f"  Listo! Abre el archivo para ver el reporte completo.")
    print(f"  {output_path}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Genera un reporte Excel con el estado de sincronización de stock."
    )
    parser.add_argument("--excel",  required=True,  help="Ruta al Excel de stock original")
    parser.add_argument("--output", required=False,
                        default=f"Reporte_Stock_Zipp_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        help="Nombre del archivo Excel de salida")
    args = parser.parse_args()
    run(excel_path=args.excel, output_path=args.output)
